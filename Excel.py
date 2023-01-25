from openpyxl.styles.borders import Border, Side
from openpyxl.utils import FORMULAE
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl import Workbook
from VLASOV_KLUCH import *
from CHEB_IST import *
from AM import *
from LUBI import *
from ZHIV_CAP import *
from MV import *
from NIAG import *
from CRYSTAL import *
from ARTVOD import *
from LUKS import *

def parsing():
    date='{:%H %M %B %Y}'.format(datetime.now())
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    fn = 'PARSER.xlsx'
    wb = load_workbook(filename='PARSER.xlsx')
    ws = wb.active
    ws = wb.create_sheet(title=str(date))
    ws.append(['', 'ВЛАСОВ КЛЮЧ', 'ЧЕБАРКУЛЬСКИЙ ИСТОК', '"КУКАЗАР" АКВАМОБИЛЬ', '"АКВА ИРЕНДЫК" АКВАМОБИЛЬ', '"ЛЕГЕНДА ГОР АРХЫЗ" АКВАМОБИЛЬ', '"СОСНОВАЯ" АКВАМОБИЛЬ', '"АРТЕНЗА" АКВАМОБИЛЬ', 'ЛЮБИМАЯ', 'ЖИВАЯ КАПЛЯ', 'ГОРНЫЙ ОАЗИС', 'NIAGARA', 'NIAGARA PREMIUM', 'NIAGARA PREMIUM С КАВКАЗСКИХ ГОР', 'КРИСТАЛЬНАЯ', '"ЛИДЕР" АРТЕЗИАНСКАЯ ВОДА' , '"ЛИДЕР PLATINUM" АРТЕЗИАНСКАЯ ВОДА', '"ЛИДЕР PROFI" АРТЕЗИАНСКАЯ ВОДА', '"ЛЮКСИК" ЛЮКС ВОДА', 'ЛЮКС ВОДА'])
    ws.append(['19л. 1 шт.'])
    ws.append(['19л. 2 шт.'])
    for i in range(1,21):
        for j in range(1,4):
            ws.cell(row=j, column=i).border = thin_border
    temp = VlKl()
    try:
        temp=int(temp)
    except:
        pass
    ws["B2"]= temp
    ws["B3"]= 0

    temp = ChIs1()
    try:
        temp=int(temp)
    except:
        pass
    ws["C2"]= temp
    temp = ChIs2()
    try:
        temp=int(temp)
    except:
        pass
    ws["C3"]= temp

    temp = Kukazar1()
    try:
        temp=int(temp)
    except:
        pass
    ws["D2"]= temp
    temp = Kukazar2()
    try:
        temp=int(temp)
    except:
        pass
    ws["D3"]= temp

    temp = Irendik1()
    try:
        temp=int(temp)
    except:
        pass
    ws["E2"]= temp
    temp = Irendik2()
    try:
        temp=int(temp)
    except:
        pass
    ws["E3"]= temp

    temp = Arhiz1()
    try:
        temp=int(temp)
    except:
        pass
    ws["F2"]= temp
    temp = Arhiz2()
    try:
        temp=int(temp)
    except:
        pass
    ws["F3"]= temp

    temp = Sosn1()
    try:
        temp=int(temp)
    except:
        pass
    ws["G2"]= temp
    temp = Sosn2()
    try:
        temp=int(temp)
    except:
        pass
    ws["G3"]= temp

    temp = Artenza1()
    try:
        temp=int(temp)
    except:
        pass
    ws["H2"]= temp
    temp = Artenza2()
    try:
        temp=int(temp)
    except:
        pass
    ws["H3"]= temp

    temp = Lubim1()
    try:
        temp=int(temp)
    except:
        pass
    ws["I2"]= temp
    temp = Lubim2()
    try:
        temp=int(temp)
    except:
        pass
    ws["I3"]= temp

    temp = Zhka1()
    try:
        temp=int(temp)
    except:
        pass
    ws["J2"]= temp
    temp = Zhka2()
    try:
        temp=int(temp)
    except:
        pass
    ws["J3"]= temp

    ws["K2"]= 0
    temp = GoOA2()
    try:
        temp=int(temp)
    except:
        pass
    ws["K3"]= temp

    temp = Ni1()
    try:
        temp=int(temp)
    except:
        pass
    ws["L2"]= temp
    temp = Ni2()
    try:
        temp=int(temp)
    except:
        pass
    ws["L3"]= temp

    temp = NiPre1()
    try:
        temp=int(temp)
    except:
        pass
    ws["M2"]= temp
    temp = NiPre2()
    try:
        temp=int(temp)
    except:
        pass
    ws["M3"]= temp

    temp = NiPreKa1()
    try:
        temp=int(temp)
    except:
        pass
    ws["N2"]= temp
    temp = NiPreKa2()
    try:
        temp=int(temp)
    except:
        pass
    ws["N3"]= temp

    temp = Cr1()
    try:
        temp=int(temp)
    except:
        pass
    ws["O2"]= temp
    temp = Cr2()
    try:
        temp=int(temp)
    except:
        pass
    ws["O3"]= temp

    temp = Ld1()
    try:
        temp=int(temp)
    except:
        pass
    ws["P2"]= temp
    temp = Ld2()
    try:
        temp=int(temp)
    except:
        pass
    ws["P3"]= temp

    temp = LdPl1()
    try:
        temp=int(temp)
    except:
        pass
    ws["Q2"]= temp
    temp = LdPl2()
    try:
        temp=int(temp)
    except:
        pass
    ws["Q3"]= temp

    temp = LdPr1()
    try:
        temp=int(temp)
    except:
        pass
    ws["R2"]= temp
    temp = LdPr2()
    try:
        temp=int(temp)
    except:
        pass
    ws["R3"]= temp

    temp = Luxic1()
    try:
        temp=int(temp)
    except:
        pass
    ws["S2"]= temp
    temp = Luxic2()
    try:
        temp=int(temp)
    except:
        pass
    ws["S3"]= temp

    temp = Lu_Vo1()
    try:
        temp=int(temp)
    except:
        pass
    ws["T2"]= temp
    temp = Lu_Vo2()
    try:
        temp=int(temp)
    except:
        pass
    ws["T3"]= temp


    wb.save(fn)


# print("=ЕСЛИ('" + TimeLast + "'!B2='" + TimeNow + "'!B2;;ABS('" + TimeLast + "'!B2-'" + TimeNow + "'!B2))")    
# ws["B4"]= ("=IF('" + TimeLast + "'!B2='" + TimeNow + "'!B2;;ABS('" + TimeLast + "'!B2-'" + TimeNow + "'!B2))")
# TimeNow = datetime